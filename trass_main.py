import math as m
from icecream import ic
from tkinter import *
from tkinter.ttk import Combobox
from tkinter import Tk, filedialog, messagebox, ttk
import datetime
import time
import scipy
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
from openpyxl import Workbook
from math_and_consts import *
import classes as cl
import dif_equations as DE
import numpy as np
from PIL import Image
import threading
#import matplotlib.image as mpimg

data = {}
data_correction = {}
ic.enable()
def calculation():
    disable_all_buttons()
    data["Rp"] = float(combo_Rp.get())
    data["e"] = float(combo_e.get())
    data["ArgLat"] = float(combo_ArgLat.get()) * cToRad
    data["Incl"] = float(combo_Incl.get()) * cToRad
    data["AscNode"] = float(combo_AscNode.get()) * cToRad
    data["ArgPerigee"] = float(combo_ArgPerig.get()) * cToRad
    data["step"] = float(combo_step.get())
    try:
        check_data(data)
    except ValueError as e:
        return
    #memo1.delete(1.0, "end")  # Clear the Text widget
    start_time = time.time()
    Calc = cl.TOrbitClass(data)
    orb = cl.TSpacecraft(data)
    orb.FCurentOrbit.assign()

    Calc.Epoch = int(datetime.datetime.now().timestamp())
    result_class_dict = Calc.class_to_cart()
    result_dict = Calc.to_geo(result_class_dict)

    try:
        s_pos = f"Положение  {result_dict['Pos']['X']:.3f}  {result_dict['Pos']['Y']:.3f}  {result_dict['Pos']['Z']:.3f}"
        memo1.insert("end", s_pos + "\n")
    except Exception as e:
        memo1.insert("end", f"Error in calculating Pos: {str(e)}\n")

    try:
        s_geo = f"Долгота  {result_dict['Longitude'] * cToDeg:.3f} (град), Широта {result_dict['Latitude'] * cToDeg:.3f} (град)"
        memo1.insert("end", s_geo + "\n")
    except Exception as e:
        memo1.insert("end", f"Error in calculating Geo coordinates: {str(e)}\n")

    try:
        s_vel = f"Скорость {result_class_dict['Vel']['X']:.3f}  {result_class_dict['Vel']['Y']:.3f}  {result_class_dict['Vel']['Z']:.3f}"
        memo1.insert("end", s_vel + "\n")
    except Exception as e:
        memo1.insert("end", f"Error in calculating Vel: {str(e)}\n")

    try:
        s_st = f"Здвездное время {Calc.SiderealTime() * cToDeg:.3f} град"
        memo1.insert("end", s_st + "\n")
    except Exception as e:
        memo1.insert("end", f"Error in calculating Sidereal Time: {str(e)}\n")
    enable_all_buttons()
    end_time = time.time()
    elapsed_time = end_time - start_time
    memo1.insert("end", f"Время расчета: {elapsed_time}" + "\n")
    memo1.insert("end", f"Текущее время: {datetime.datetime.now()}" + "\n")
    memo1.insert("end", f"Время запуска спутника совпадает с текущим временем \n")


def graf():
    disable_all_buttons()
    ic.enable()
    data["Rp"] = float(combo_Rp.get())
    data["e"] = float(combo_e.get())
    data["ArgLat"] = float(combo_ArgLat.get()) * cToRad
    data["Incl"] = float(combo_Incl.get()) * cToRad
    data["AscNode"] = float(combo_AscNode.get()) * cToRad
    data["ArgPerigee"] = float(combo_ArgPerig.get()) * cToRad
    data["step"] = float(combo_step.get())
    data["interval"] = float(combo_interval.get())
    try:
        check_data(data)
    except ValueError as e:
        return
    start_time = time.time()
    calc = cl.TSpacecraft(data)
    orb = cl.TOrbitClass(data)
    my_time = datetime.datetime.now().timestamp()
    end_time = my_time + data["interval"]
    time_step = orb.step
    x = []
    y = []
    counter = 0
    while my_time < end_time:
        result_class_dict_graf, result_dict_graf = calc.change_time(my_time)
        x_values = result_dict_graf['Longitude'] * cToDeg
        x.append(x_values)
        y_values = result_dict_graf['Latitude'] * cToDeg
        y.append(y_values)
        my_time += time_step
        counter += 1
        ic(my_time)
    end_time = time.time()
    elapsed_time = end_time - start_time
    memo1.insert("end", f"Время работы graf: {elapsed_time} секунд\n")
    memo1.insert("end", f"Количество итераций:{counter}\n")

    map_image_path = r"C:\Users\zheny\OneDrive\Рабочий стол\xxxxxx\универ\IT\питон самоучение\spacecraft\images.jpg"
    map_image = Image.open(map_image_path)
    map_image = map_image.convert("RGBA")  # Конвертируем изображение в формат RGBA
    map_image = np.array(map_image)  # Преобразуем изображение в массив numpy

    window_2 = Tk()
    window_2.title('Трасса')

    # Размеры окна
    window_width = 700
    window_height = 500
    window_2.geometry(f"{window_width}x{window_height}")

    # Создаем главный фрейм для хранения графиков и скроллбара
    plot_width = window_width * 0.9  # Ширина одного графика
    plot_height = int(plot_width * 0.66)  # Высота одного графика

    # Создаем фигуру Matplotlib
    fig, ax = plt.subplots(figsize=(plot_width / 100, plot_height / 100), dpi=100)
    ax.scatter(x, y)  # Отображаем соответствующий график

    # Отображаем изображение на фоне
    ax.imshow(map_image, extent=[-180, 180, -90, 90], aspect='auto')
    # Настройки графика (добавьте необходимые настройки)
    ax.set_title("Трасса спутника")
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.set_xlim(-180, 180)
    ax.set_ylim(-90, 90)

    # Создаем объект FigureCanvasTkAgg для отображения графика в Tkinter
    canvas = FigureCanvasTkAgg(fig, master=window_2)
    canvas.draw()

    # Размещаем объект FigureCanvasTkAgg на внутреннем фрейме
    canvas.get_tk_widget().pack(side=TOP, pady=10, padx=30)

    # Запуск основного цикла обработки событий
    window_2.mainloop()
    enable_all_buttons()

def correction():
    disable_all_buttons()
    start_time = time.time()
    data["Rp"] = float(combo_Rp.get())
    data["e"] = float(combo_e.get())
    data["ArgLat"] = float(combo_ArgLat.get()) * cToRad
    data["Incl"] = float(combo_Incl.get()) * cToRad
    data["AscNode"] = float(combo_AscNode.get()) * cToRad
    data["ArgPerigee"] = float(combo_ArgPerig.get()) * cToRad
    data["step"] = float(combo_step.get())
    data["thrust"] = float(combo_thrust.get())
    data["integr"] = str(combo_integr.get())
    data["interval"] = float(combo_interval.get())
    data["work"] = float(combo_work.get())
    data["turn"] = float(combo_turn.get())
    data["stabelize"] = str(combo_stabelize.get())
    data["mass"] = int(combo_mass.get())
    data["direction"] = str(combo_direction.get())

    try:
        check_data(data)
    except ValueError as e:
        return

    data_correction = data.copy()
    eq = DE.Difur(data)
    orb = cl.TOrbitClass(data)
    calc = cl.TSpacecraft(data_correction)

    if float(combo_interval.get()) == 0:
        data["interval"] = orb.period()
    my_time = 0
    time_step = orb.step
    X = []; A =[]; P = []; E = []; R = []; OM = []; Lon = []; Lat = []
    e, omega, atta, p = data["e"], data["ArgPerigee"], orb.true_anomaly(), orb.parameter()

    if atta > 2 * m.pi:
        atta %= (2 * m.pi)
    if atta < 0:
        atta += 2 * m.pi

    counter = 0
    r = (p / (1 + e * m.cos(atta)))
    initial = data.copy()
    initial["omega"] = omega
    initial["atta"] = atta
    initial["p"] = p
    initial["r"] = r

    stabilization_algorithms = {
        "Постоянство эксцентриситета": eq.stabilization_esccentr,
        "Максимальная скорость изменения эксцентриситета": eq.stabilization_max_speed_esccentr,
        "Постоянство расстояния до перигея": eq.stabilization_perig,
        "Максимальная скорость изменения расстояния до перигея": eq.stabilization_max_speed_perig,
        "Постоянство расстояния до апогея": eq.stabilization_apog,
        "Максимальная скорость изменения расстояния до апогея": eq.stabilization_max_speed_apog,
        "Постоянство фокального параметра": eq.stabilization_phocpar,
        "Максимальная скорость изменения фокального параметра": eq.stabilization_max_speed_phocpar,
        "Постоянство большой полуоси": eq.stabilization_a,
        "Максимальная скорость изменения большой полуоси": eq.stabilization_max_speed_a,
        "Постоянство положения линии аписд": eq.stabilization_apsid,
        "Максимальная скорость вращения линии аписд": eq.stabilization_max_speed_apsid
    }

    intearated_methods = {
        "Эйлера": eq.euler,
        "Эйлера-Коши": eq.euler_cauchy,
        "Рунге-Кутты 4": eq.runge_kutta_4,
    }

    equations = [eq.de_func, eq.dp_func, eq.domega_func, eq.datta_func]

    while my_time < data["interval"]:

        if data["turn"] <= my_time <= (data["turn"] + data["work"]):
            F = data["thrust"]/data["mass"]/1000

            if data["stabelize"] in stabilization_algorithms:
                lam = stabilization_algorithms[data["stabelize"]](e, atta)

        elif not (data["turn"] <= my_time <= (data["turn"] + data["work"])):
            lam = 0
            F = 0

        if data["direction"] == "Отрицательное":
            F = -F

        initial["F"] = F
        initial["lam"] = lam
        dx =['e', 'p', 'omega', 'atta']
        if data["integr"] in intearated_methods:
            values = intearated_methods[data["integr"]](equations, initial, data["step"], dx)

        e = values[0]
        p = values[1]
        omega = values[2]
        atta = values[3]

        if atta > 2 * m.pi:
            atta %= (2 * m.pi)
        if atta < 0:
            atta += 2 * m.pi

        arglat = omega + atta
        if arglat > 2 * m.pi:
            arglat %= (2 * m.pi)
        if arglat < 0:
            arglat += 2 * m.pi
        r = p / (1 + e * m.cos(atta))
        a = p / (1 - e ** 2)

        #еще ода вариация из одной и той же методички (результаты обе формулы дают одинаковые)
        #a = (cMu / ((2 * m.pi) / (2 * m.pi * (p / (1 - e ** 2)) * m.sqrt((p / (1 - e ** 2)) / cMu))) ** 2) ** (1 / 3)
        initial["e"] = e
        initial["omega"] = omega
        initial["atta"] = atta
        initial["p"] = p
        initial["r"] = r
        A.append(a); P.append(p); E.append(e); R.append(r); OM.append(omega * cToDeg); X.append(my_time)

        data_correction["Rp"] = p / (1 + e)
        data_correction["e"] = e
        data_correction["ArgLat"] = arglat #аргумент широты
        data_correction["Incl"] = float(combo_Incl.get()) * cToRad
        data_correction["AscNode"] = float(combo_AscNode.get()) * cToRad #долгоа восходящего узла
        data_correction["ArgPerigee"] = omega
        data_correction["step"] = float(combo_step.get())
        result_class_dict_graf, result_dict_graf = calc.update(data_correction, my_time)
        Longitude = result_dict_graf['Longitude'] * cToDeg
        Latitude = result_dict_graf['Latitude'] * cToDeg
        Lon.append(Longitude); Lat.append(Latitude)
        my_time += time_step
        counter += 1

    enable_all_buttons()
    end_time = time.time()
    elapsed_time = end_time - start_time
    memo1.insert("end", f"Время работы graf: {elapsed_time} секунд\n")
    memo1.insert("end", f"Количество итераций:{counter}\n")
    save_to_excel(A, P, E, R, OM, X)
    plot_graphs_with_scrollbar(X, A, P, E, R, OM, Lon, Lat)


def _on_mouse_wheel(event, canvas):
    """Обработчик событий для прокрутки колесика мыши."""
    canvas.yview_scroll(int(-1 * (event.delta / 100)), "units")


def plot_graphs_with_scrollbar(X, A, P, E, R, OM, Lon, Lat):
    # Создаем главное окно Tkinter
    window = Tk()
    window.title('Графики')

    # Размеры окна
    window_width = 750
    window_height = 600
    window.geometry(f"{window_width}x{window_height}")

    # Создаем главный фрейм для хранения графиков и скроллбара
    main_frame = Frame(window)
    main_frame.pack(fill=BOTH, expand=1)

    # Создаем Canvas для отображения графиков и привязываем к нему скроллбар
    canvas_graf = Canvas(main_frame)
    canvas_graf.pack(side=LEFT, fill=BOTH, expand=1)
    scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=canvas_graf.yview)
    scrollbar.pack(side=RIGHT, fill=Y)
    canvas_graf.configure(yscrollcommand=scrollbar.set)
    canvas_graf.bind("<Configure>", lambda e: canvas_graf.configure(scrollregion=canvas_graf.bbox("all")))

    # Создаем внутренний фрейм для размещения графиков
    inner_frame = Frame(canvas_graf)
    canvas_graf.create_window((0, 0), window=inner_frame, anchor="nw")

    # Создаем и отображаем графики
    num_plots = 5
    plot_width = window_width * 0.9  # Ширина одного графика
    plot_height = int(plot_width * 0.66)  # Высота одного графика

    for i in range(num_plots):
        # Создаем фигуру Matplotlib
        fig, ax = plt.subplots(figsize=(plot_width / 100, plot_height / 100), dpi=100)
        ax.scatter(X, [A, P, E, R, OM][i])  # Отображаем соответствующий график

        # Настройки графика
        ax.set_title(['Большая полуось', 'Фокальный параметр', 'Эксцентриситет', 'Радиус орбиты', 'Аргумент перигея'][i])
        ax.grid(True)

        # Создаем объект FigureCanvasTkAgg для отображения графика в Tkinter
        canvas = FigureCanvasTkAgg(fig, master=inner_frame)
        canvas.draw()

        # Размещаем объект FigureCanvasTkAgg на внутреннем фрейме
        canvas.get_tk_widget().pack(side=TOP, pady=10, padx=30)

    map_image_path = r"C:\Users\zheny\OneDrive\Рабочий стол\xxxxxx\универ\IT\питон самоучение\spacecraft\images.jpg"
    map_image = Image.open(map_image_path)
    map_image = map_image.convert("RGBA")  # Конвертируем изображение в формат RGBA
    map_image = np.array(map_image)  # Преобразуем изображение в массив numpy
    # Создаем фигуру Matplotlib
    fig, ax = plt.subplots(figsize=(plot_width / 100, plot_height / 100), dpi=100)
    ax.scatter(Lon, Lat)  # Отображаем соответствующий график

    # Отображаем изображение на фоне
    ax.imshow(map_image, extent=[-180, 180, -90, 90], aspect='auto')
    # Настройки графика (добавьте необходимые настройки)
    ax.set_title("Трасса спутника")
    ax.set_xlabel("X")
    ax.set_ylabel("Y")
    ax.set_xlim(-180, 180)
    ax.set_ylim(-90, 90)

    # Создаем объект FigureCanvasTkAgg для отображения графика в Tkinter
    canvas = FigureCanvasTkAgg(fig, master=inner_frame)
    canvas.draw()

    # Размещаем объект FigureCanvasTkAgg на внутреннем фрейме
    canvas.get_tk_widget().pack(side=TOP, pady=10, padx=30)
    window.bind_all("<MouseWheel>", mouse_wheel)
    window.mainloop()



def load_from_file():
    file_path = filedialog.askopenfilename(filetypes=[("Text files", "*.txt")])

    if file_path:
        with open(file_path, "r") as file:
            lines = file.readlines()

        # Создаем словарь для хранения значений
        values_dict = {}

        # Разбираем строки файла
        for line in lines:
            # Разбиваем строку по знаку '='
            parts = line.split('=')

            # Если не удалось разделить строку на две части, пропускаем ее
            if len(parts) != 2:
                continue

            # Получаем ключ и значение
            key = parts[0].strip()
            value = parts[1].strip()

            values_dict[key] = value

        # Устанавливаем значения в комбобоксы
        combo_Rp.set(values_dict.get('Prigee', ''))
        combo_e.set(values_dict.get('Eccentr', ''))
        combo_ArgLat.set(values_dict.get('ArgLat', ''))
        combo_ArgPerig.set(values_dict.get('PeriArg', ''))
        combo_AscNode.set(values_dict.get('AscNode', ''))
        combo_Incl.set(values_dict.get('Inclin', ''))
        combo_step.set(values_dict.get('Step', ''))
        combo_mass.set(values_dict.get('mass', ''))
        combo_thrust.set(values_dict.get('thrust', ''))
        combo_integr.set(values_dict.get('Integr', ''))
        combo_interval.set(values_dict.get('interval', ''))
        combo_work.set(values_dict.get('work', ''))
        combo_turn.set(values_dict.get('turn', ''))
        combo_stabelize.set(values_dict.get('stabelize', ''))
        combo_direction.set(values_dict.get('direction', ''))
def save_to_file():
    data = f"Аргумент перигея Perigee={combo_Rp.get()}\n" \
           f"Эксцентриситет Eccentr={combo_e.get()}\n" \
           f"Аргумент широты ArgLat={combo_ArgLat.get()}\n" \
           f"Аргумент перигея PeriArg={combo_ArgPerig.get()}\n" \
           f"Долгота восходящего узла RAAN={combo_AscNode.get()}\n" \
           f"Наклонение орбиты Inclin={combo_Incl.get()}\n" \
           f"Дата выполнения операции Date={datetime.datetime.now().date()}\n" \
           f"Время выполнения операции Time={datetime.datetime.now().time()}\n" \
           f"Шаг моделирования step={combo_step.get()}\n"\
           f"Метод интегрирования integr={combo_integr.get()}\n"\
           f"Интервал моделирования interval={combo_interval.get()}\n"\
           f"Время работы двигателя work={combo_work.get()}\n"\
           f"Время включения двигателя turn={combo_turn.get()}\n"\
           f"Алгоритм стабилизации stabelize={combo_stabelize.get()}\n"\
           f"Направления тяги двигателя direction={combo_direction.get()}\n"\
           f"\nMemo Text:\n{memo1.get('1.0', 'end')}"


    file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text files", "*.txt")])

    if file_path:
        with open(file_path, "w") as file:
            file.write(data)

def save_to_excel(A, P, E, R, OM, X, step = 10):
    # Создаем новый рабочий лист
    wb = Workbook()  # Создаем новый объект Workbook от openpyxl
    ws = wb.active  # Делаем активным первый (и единственный) лист в новой книге
    ws.title = "Satellite Data"  # Устанавливаем название листа

    # Заголовки для параметров
    data = [  # Создаем список кортежей, каждый из которых содержит имя параметра и его значение
        ("Аргумент перигея", combo_Rp.get()),
        ("Эксцентриситет", combo_e.get()),
        ("Аргумент широты", combo_ArgLat.get()),
        ("Аргумент перигея", combo_ArgPerig.get()),
        ("Долгота восходящего узла", combo_AscNode.get()),
        ("Наклонение орбиты", combo_Incl.get()),
        ("Текущая дата и время", datetime.datetime.now().date()),
        ("Время выполнения операции", datetime.datetime.now().time()),
        ("Шаг моделирования", combo_step.get()),
        ("Метод интегрирования", combo_integr.get()),
        ("Интервал моделирования", combo_interval.get()),
        ("Время работы двигателя", combo_work.get()),
        ("Время включения двигателя", combo_turn.get()),
        ("Алгоритм стабилизации", combo_stabelize.get()),
        ("Направление тяги двигателя", combo_direction.get())
    ]

    # Записываем параметры в первые строки
    for i, (param, value) in enumerate(data, start=1):
        ws[f'A{i}'] = param  # Записываем имя параметра в колонку A
        ws[f'B{i}'] = value  # Записываем значение параметра в колонку B

    # Заголовки для массивов данных
    headers = ["Время", "a", "p", "e", "r", "omega"]  # Список заголовков для данных
    for col_num, header in enumerate(headers, start=1):
        ws.cell(row=len(data) + 2, column=col_num, value=header)  # Записываем заголовки в строку, следующую за параметрами

    # Записываем данные массивов в столбцы с заданным шагом
    max_len = max(len(X), len(A), len(P), len(E), len(R), len(OM))  # Находим максимальную длину среди массивов данных
    start_row = len(data) + 3  # Начинаем с первой строки после заголовков

    for row_num in range(0, max_len, step):
        if row_num < len(X):
            ws.cell(row=start_row + row_num // step, column=1, value=X[row_num])  # Записываем данные времени
        if row_num < len(A):
            ws.cell(row=start_row + row_num // step, column=2, value=A[row_num])  # Записываем данные a
        if row_num < len(P):
            ws.cell(row=start_row + row_num // step, column=3, value=P[row_num])  # Записываем данные p
        if row_num < len(E):
            ws.cell(row=start_row + row_num // step, column=4, value=E[row_num])  # Записываем данные e
        if row_num < len(R):
            ws.cell(row=start_row + row_num // step, column=5, value=R[row_num])  # Записываем данные r
        if row_num < len(OM):
            ws.cell(row=start_row + row_num // step, column=6, value=OM[row_num])  # Записываем данные omega

    '''#Добавляем текст из memo, если надо взять значения без шага, а сразу все 
    memo_text = memo1.get('1.0', 'end')  # Получаем текст из memo
    memo_start_row = start_row + (max_len // step) + 2  # Определяем начальную строку для memo
    ws.cell(row=memo_start_row, column=1, value="Memo Text")  # Пишем заголовок для memo
    for i, line in enumerate(memo_text.splitlines(), start=memo_start_row + 1):
        ws.cell(row=i, column=1, value=line)  # Записываем каждую строку memo в новую строку таблицы'''

    # Выбираем путь для сохранения файла
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel files", "*.xlsx")])
    # Открывается диалоговое окно для сохранения файла с предложением сохранить с расширением .xlsx

    if file_path:
        wb.save(file_path)  # Если путь выбран, сохраняем файл по указанному пути

def check_data(data):
    # Определяем допустимые диапазоны значений для каждого параметра
    valid_ranges = {
        "Rp": (7000, 43000),           # Например, радиус перицентра должен быть положительным числом
        "e": (0, 0.8),                       # Эксцентриситет должен быть в диапазоне от 0 до 1
        "ArgLat": (0, 360),                # Аргумент широты должен быть в диапазоне от 0 до 360 градусов
        "Incl": (-90, 90),                  # Наклонение должно быть в диапазоне от 0 до 180 градусов
        "AscNode": (0, 360),               # Долгота восходящего узла должна быть в диапазоне от 0 до 360 градусов
        "ArgPerigee": (0, 360),            # Аргумент перигея должен быть в диапазоне от 0 до 360 градусов
    }

    # Проверяем наличие всех необходимых ключей в словаре
    for key in valid_ranges:
        if key not in data:
            messagebox.showerror("Ошибка", f"Отсутствует ключ '{key}' в данных")
            raise ValueError(f"Отсутствует ключ '{key}' в данных")

        # Проверяем значения каждого ключа на допустимость
    for key, (min_val, max_val) in valid_ranges.items():
        value = data[key]
        if not isinstance(value, (int, float)):
            messagebox.showerror("Ошибка", f"Значение '{key}' должно быть числом")
            raise ValueError(f"Значение '{key}' должно быть числом")
        if not (min_val <= value <= max_val):
            messagebox.showerror("Ошибка",
                                 f"Значение '{key}' ({value}) вне допустимого диапазона {min_val} - {max_val}")
            raise ValueError(f"Значение '{key}' ({value}) вне допустимого диапазона {min_val} - {max_val}")

    return True

def mouse_wheel(event):
    # Проверяем направление прокрутки
    if event.delta > 0:
        my_canvas.yview_scroll(-1, "units")
    else:
        my_canvas.yview_scroll(1, "units")

root = Tk()

root.title(" Алтухов Е.С. М6О-201С-22")
root.geometry("800x800")

main_frame = Frame(root)
main_frame.pack(fill=BOTH, expand=1)
my_canvas = Canvas(main_frame)
my_canvas.pack(side=LEFT, fill=BOTH, expand=1)
my_scrollbar = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
my_scrollbar.pack(side=RIGHT, fill=Y)
my_canvas.configure(yscrollcommand=my_scrollbar.set)
my_canvas.bind("<Configure>", lambda e: my_canvas.configure(scrollregion=my_canvas.bbox("all")))
second_frame = Frame(my_canvas)
my_canvas.create_window((0, 0), window=second_frame, anchor="nw")

# Привязываем события прокрутки к канвасу
my_canvas.bind_all("<MouseWheel>", mouse_wheel)  # Для Windows

label_Rp = Label(second_frame, text="Введите большую полуось", font=("Times New Roman", 12), fg="blue")
label_Rp.pack()
combo_Rp = Combobox(second_frame)
combo_Rp['values'] = (8000.0, 20_000.0, 42_000.0, "Свое значение")
combo_Rp.current(0)
combo_Rp.pack(padx=330)

label_e = Label(second_frame, text="Введите эксцентриситет", font=("Times New Roman", 12), fg="blue")
label_e.pack()
combo_e = Combobox(second_frame)
combo_e['values'] = (0.6, 0.05, 0.03, "Свое значение")
combo_e.current(2)
combo_e.pack()

label_ArgLat = Label(second_frame, text="Введите Аргумент широты, град ", font=("Times New Roman", 12), fg="blue")
label_ArgLat.pack()
combo_ArgLat = Combobox(second_frame)
combo_ArgLat['values'] = (90.0, 60.0, 20.0, "Свое значение")
combo_ArgLat.current(0)
combo_ArgLat.pack()

label_ArgPerig = Label(second_frame, text="Введите аргумент Перигея, град", font=("Times New Roman", 12), fg="blue")
label_ArgPerig.pack()
combo_ArgPerig = Combobox(second_frame)
combo_ArgPerig['values'] = (60.0, 270.0, 90.0, "Свое значение")
combo_ArgPerig.current(1)
combo_ArgPerig.pack()

label_AscNode = Label(second_frame, text="Введите долготу восходящего узла, град", font=("Times New Roman", 12), fg="blue")
label_AscNode.pack()
combo_AscNode = Combobox(second_frame)
combo_AscNode['values'] = (30.0, 90.0, 200.0, "Свое значение")
combo_AscNode.current(1)
combo_AscNode.pack()

label_Incl = Label(second_frame, text="Введите Наклонение орбиты, град", font=("Times New Roman", 12), fg="blue")
label_Incl.pack()
combo_Incl = Combobox(second_frame)
combo_Incl['values'] = (10, 20, 60, "Свое значение")
combo_Incl.current(1)
combo_Incl.pack()

label_step = Label(second_frame, text="Введите шаг моделирования графика, секунд", font=("Times New Roman", 12), fg="blue")
label_step.pack()
combo_step = Combobox(second_frame)
combo_step['values'] = (20, 30, 60, "Свое значение")
combo_step.current(0)
combo_step.pack()

label_mass = Label(second_frame, text="Введите массу КА", font=("Times New Roman", 12), fg="blue")
label_mass.pack()
combo_mass = Combobox(second_frame)
combo_mass['values'] = (500, 3000, 3500, "Свое значение")
combo_mass.current(2)
combo_mass.pack()

label_thrust = Label(second_frame, text="Введите тягу двигателя", font=("Times New Roman", 12), fg="blue")
label_thrust.pack()
combo_thrust = Combobox(second_frame)
combo_thrust['values'] = (1, 5, 10, "Свое значение")
combo_thrust.current(2)
combo_thrust.pack()

label_integr = Label(second_frame, text="Выберете метод интегрирования", font=("Times New Roman", 12), fg="blue")
label_integr.pack()
combo_integr = Combobox(second_frame)
combo_integr['values'] = ("Эйлера", "Эйлера-Коши", "Рунге-Кутты 4")
combo_integr.current(2)
combo_integr.pack()

label_interval = Label(second_frame, text="Введите интервал интегрирования, с", font=("Times New Roman", 12), fg="blue")
label_interval.pack()
combo_interval = Combobox(second_frame)
combo_interval['values'] = (84325, 55000, 7200, "Свое значение")
combo_interval.current(2)
combo_interval.pack()

label_work = Label(second_frame, text="Введите время работы двигателя", font=("Times New Roman", 12), fg="blue")
label_work.pack()
combo_work = Combobox(second_frame)
combo_work['values'] = (100, 500, 150, "Свое значение")
combo_work.current(1)
combo_work.pack()

label_turn = Label(second_frame, text="Введите время включения двигателя", font=("Times New Roman", 12), fg="blue")
label_turn.pack()
combo_turn = Combobox(second_frame)
combo_turn['values'] = (2000, 50, 150, "Свое значение")
combo_turn.current(0)
combo_turn.pack()

label_stabelize = Label(second_frame, text="Выберите алгоритм стабилизации", font=("Times New Roman", 12), fg="blue")
label_stabelize.pack()
combo_stabelize = Combobox(second_frame)
combo_stabelize['values'] = ("Постоянство эксцентриситета", "Максимальная скорость изменения эксцентриситета",
                        "Постоянство расстояния до перигея", "Максимальная скорость изменения расстояния до перигея",
                        "Постоянство расстояния до апогея", "Максимальная скорость изменения расстояния до апогея",
                        "Постоянство фокального параметра", "Максимальная скорость изменения фокального параметра",
                        "Постоянство большой полуоси", "Максимальная скорость изменения большой полуоси",
                        "Постоянство положения линии аписд", "Максимальная скорость вращения линии аписд")
combo_stabelize.current(0)
combo_stabelize.pack()

label_direction = Label(second_frame, text="Выберите направление тяги", font=("Times New Roman", 12), fg="blue")
label_direction.pack()
combo_direction = Combobox(second_frame)
combo_direction['values'] = ("Положительное", "Отрицательное")
combo_direction.current(0)
combo_direction.pack()

def graf_threat():
    waiting = f"Подождите, это займет какое-то время"
    memo1.insert("end", waiting + "\n")
    thread = threading.Thread(target=graf)
    thread.start()

def correction_thread():
    waiting = f"Подождите, это займет какое-то время"
    memo1.insert("end", waiting + "\n")
    thread2 = threading.Thread(target=correction)
    thread2.start()

def disable_all_buttons():
    btn_load.config(state="disabled")
    btn_calc.config(state="disabled")
    btn_graf.config(state="disabled")
    btn_correction.config(state="disabled")
    btn_save.config(state="disabled")

def enable_all_buttons():
    btn_load.config(state="normal")
    btn_calc.config(state="normal")
    btn_graf.config(state="normal")
    btn_correction.config(state="normal")
    btn_save.config(state="normal")

btn_load = Button(root, text="Загрузить данные из файла", font=("Times New Roman", 12, "bold"), fg="blue", command=load_from_file)
btn_load.pack()

btn_calc = Button(root, text="Расчет", font=("Times New Roman", 12, "bold"), fg="red", command=calculation)
btn_calc.pack()

btn_graf = Button(root, text="Построить трассу спутника", font=("Times New Roman", 12, "bold"), fg="red", command=graf_threat)
btn_graf.pack()

btn_correction = Button(root, text="Смоделировать движение по орбите", font=("Times New Roman", 12, "bold"), fg="red", command=correction_thread)
btn_correction.pack()

memo1 = Text(root, wrap="none", height=20, width=90)
memo1.pack()

btn_save = Button(root, text="Сохранить в блокнот", font=("Times New Roman", 12, "bold"), fg="red", command=save_to_file)
btn_save.pack()

root.mainloop()




